from openpyxl import load_workbook


#
# load excel sheet and hide unneeded columns and rows and create new file called
# cleanedinventorycount.xlsx
#

# loads excel file
wb = load_workbook(filename = 'workbook.xlsx')
ws = wb.active
# ask the end user range of rows needed
x = int(raw_input("please input row number of the first phone:  "))
y = int(raw_input("please input row number of the last phone: "))
# hides rest of the rows user has not selected
for toprow in range (8,x-1):
    ws.row_dimensions[toprow].hidden = True

for botrow in range (y+1,ws.get_highest_row()):
    ws.row_dimensions[botrow].hidden = True

for celcolumn in range(x,y+1):
    if ws.cell(row = celcolumn, column = 4).value == 0:
        ws.row_dimensions[celcolumn].hidden = True

# create new file
wb.save('cleanedinventorycount.xlsx')





